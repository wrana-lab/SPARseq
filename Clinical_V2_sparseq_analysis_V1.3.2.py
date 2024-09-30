#!/usr/bin/python

import os, sys
import argparse
import re
import openpyxl
from openpyxl import Workbook

# Last updated: Sep 20th 2021 {KRZ}
# Inputs:
#       select_vars - *.fastq files
#       all_top - *.fastq files and codon_aa_table.csv
#       bowtie_count - *sam.count.txt files
#       variant_agg - runcl*_CountTable.txt, SparSeq_Srbd_v2_top1_Var_summary_table.txt and runcl*_SelectedVariants.txt files
# Ouputs:
#       select_vars - R1R2_Srbd_v2_top_PCR_hits.txt, R1R2_Spbs_top_PCR_hits.txt, R1R2_RdRP_top_PCR_hits.txt, important_vars_detailed_vX.txt, runclXX_SelectedVariants.xlsx
#       all_top - SparSeq_RdRP_top1_Var_summary_table.txt, SparSeq_Srbd_v2_top1_Var_summary_table.txt, SparSeq_Spbs_top1_Var_summary_table.txt, runclXX_AllVariantDetails.xlsx
#       bowtie_count - bowtie_count_table_v#.txt, runclXX_CountTable.txt, runclXX_CountTable.xlsx
#       variant_agg - sparseq_report_runclXX.xlsx
#
# Versioning:
#    V1.3
#      - See "SPAR_Seq Pipeline Wrapper Info V1.3" for full details
#    V1.3.1
#       - Updated script version regex to account for 'V'
#       - Outputs numbers instead of text for runclXX_CountTable.xlsx
#    V1.3.2
#       - Fixed issue with regex via increasing amount of allowed sample digits from {1,2} to {1,4}

def main():

    parser = argparse.ArgumentParser(description = 'This outputs an alignment file for each amplicon and \'important_vars_detailed_v3.txt\'. This important_vars file is what quickly tells us if there is a variant')
    parser.add_argument('run_folder', help = 'Full path of run folder with fastq list file')
    parser.add_argument('resource_folder', help = 'Full path of SPAR Seq resource folder')
    parser.add_argument('--start', help='Stage to start in [select_vars|all_top|bowtie_count|variant_agg] Default: select_vars')
    parser.add_argument('--intermediates', help = 'Keep Intermediates [Y|N] DEFAULT: N') #Optional argument: --intermediates option
    args = parser.parse_args()  #Needed for args to be recognized

    args.run_folder = os.path.abspath(args.run_folder) #abspath
    args.resource_folder = os.path.abspath(args.resource_folder)

    # Debug options: 1 == true, 0 == false
    run_select_vars = 1
    run_all_top = 1
    run_bowtie_count = 1
    run_variant_agg = 1

    if args.start:
        print("\t\t Analysis pipeline starting at " + args.start + ":")
        if args.start == 'all_top':
            run_select_vars = 0
        elif args.start == 'bowtie_count':
            run_select_vars = 0
            run_all_top = 0
        elif args.start == 'variant_agg':
                run_select_vars = 0
                run_all_top = 0
                run_bowtie_count = 0
    else:
        print("\t\t Analysis pipeline starting at select_vars:")

    #parsing script version
    versionre = re.match(r'.*/.[^/]*(V.+)\.py', sys.argv[0])
    version = versionre.group(1)

    #runID
    runIDre = re.match(r'.*(runcl.[^/]*).*', sys.argv[1])
    runID = runIDre.group(1)

    #script
    scriptre = re.match(r'.*/(.[^/]+)\.py', sys.argv[0])
    script = scriptre.group(1)


    sampleID_list = [] #used to check for duplicate sampleIDs
    input_file_set = []

    for fastq in os.listdir(args.run_folder+"/R1_files"):
        if "_R1_" in fastq and ".fastq" in fastq and ".sam" not in fastq:
            matches = re.match(r'.*?_(.*?)_.*?', fastq)
            sampleID_list.append(matches.group(1))
            input_file_set.append(fastq)

    input_file_set.sort()

    #create list of duplicate samples
    sID = set()
    sampleID_dups = {x for x in sampleID_list if x in sID or (sID.add(x) or False)}

    if run_select_vars or run_all_top:

        #ROI
        RdRP_amp="GATGCCACAACTGCTTATGCTAATAGTGTTTTTAACATTTGTCAAGCTGTCACGGCCAATGTTAATGCACTTTTATCTACTGATGGTAACAAAATTGCCGATAAGTATGTCCGCAA"
        Srbd_v2_amp="ACCTTTTGAGAGAGATATTTCAACTGAAATCTATCAGGCCGGTAGCACACCTTGTAATGGTGTTGAAGGTTTTAATTGTTACTTTCCTTTACAATCATATGGTTTCCAACCCACTAATGGTGTTGGTTACCAACCATACAGAGTAGTAGT"
        Spbs_amp="TATGCGCTAGTTATCAGACTCAGACTAATTCTCCTCGGCGGGCACGTAGTGTAGCTAGTCAATCCATCATTGCCTACACTATGTCACTTGGTGCAGAAAATTCAGTTGCTTAC"

        #define select_var output files
        if run_select_vars:

            Srbd_v2_PCR_path = args.run_folder+"/R1R2_Srbd_v2_top_PCR_hits.txt"
            Srbd_v2_PCR_file = open(Srbd_v2_PCR_path, "w")
            Srbd_v2_PCR_file.write(">Refseq_Srbd_v2_pst_strand" + "\n" + Srbd_v2_amp + "\n")

            Spbs_PCR_path = args.run_folder+"/R1R2_Spbs_top_PCR_hits.txt"
            Spbs_PCR_file = open(Spbs_PCR_path, "w")
            Spbs_PCR_file.write(">Refseq_Spbs_pst_strand" + "\n" + Spbs_amp + "\n")

            RdRP_PCR_path = args.run_folder+"/R1R2_RdRP_top_PCR_hits.txt"
            RdRP_PCR_file = open(RdRP_PCR_path, "w")
            RdRP_PCR_file.write(">Refseq_RdRP_pst_strand" + "\n" + RdRP_amp + "\n")

            fileout_combMutpath = args.run_folder+"/results/important_vars_detailed_"+version+".txt"
            fileout_combMut = open(fileout_combMutpath, "w")
            second_info="sample,Srbd_v2_tophit,Spbs_tophit,Srbd_v2_tophit_percentage,Srbd_v2_total_read_count,Spbs_tophit_percentage,Spbs_total_read_count,Srbd_v2_WT_percentage,total_N501Y_percentage,total_E484K_percentage,total_S494P_percentage,Spbs_WT_percentage,total_P681H_percentage,total_P681R_percentage,total_T478K_percentage,file"
            second_info_l = second_info.split(",")
            fileout_combMut.write(second_info + "\n")

            selvarstablepath = args.run_folder+"/"+runID+"_SelectedVariants.txt"
            selvarstable = open(selvarstablepath, "w")
            selvarstable.write(second_info+"\n")

            #generating workbook in parallel
            selvars_wbname = args.run_folder+"/results/"+runID+"_SelectedVariants.xlsx"
            selvars_wb = Workbook()
            ws_selvars = selvars_wb.active
            ws_selvars.title = "important_vars_detailed_"+version
            ws_selvars.append(second_info_l)#adding header to wb

            second_info_l.pop() #remove filename to add additional fields inbetween

        #define select_var output files
        if run_all_top:
            RdRP_Var_path = args.run_folder+"/results/SparSeq_RdRP_top1_Var_summary_table.txt"
            RdRP_Var_file = open(RdRP_Var_path, "w")
            RdRP_Var_file.write("sample,well,RdRP_total_read#,RdRP_top1_read#,nt_var,aa_var,present_in_sec_hit,present_in_third_hit,filename" + "\n")

            Srbd_v2_Var_path = args.run_folder+"/results/SparSeq_Srbd_v2_top1_Var_summary_table.txt"
            Srbd_v2_Var_file = open(Srbd_v2_Var_path, "w")
            Srbd_v2_Var_file.write("sample,well,Srbd_v2_total_read#,Srbd_v2_top1_read#,nt_var,aa_var,present_in_sec_hit,present_in_third_hit,filename" + "\n")

            Spbs_Var_path = args.run_folder+"/results/SparSeq_Spbs_top1_Var_summary_table.txt"
            Spbs_Var_file = open(Spbs_Var_path, "w")
            Spbs_Var_file.write("sample,well,Spbs_total_read#,Spbs_top1_read#,nt_var,aa_var,present_in_sec_hit,present_in_third_hit,filename" + "\n")
            #fileout.write("sample info,variation(s) in top-seq, top seq variant(S) present in 2.top-hit?,top seq variant(S) present in 3.top-hit?, top-hit read count, variant(s) in 2.top-hit, 2.top-hit seq variant(S) present in 3.top-hit?, 2.top-hit read count, variants in 3.top-hit, 3.top-hit read count, total amplicon read count" + "\n" )

            #fileout1 = open("RdRP_nt_list.txt", "w")
            #fileout2 = open("RdRP_aa_list.txt", "w")
            #fileout3 = open("Srbd_v2_nt_list.txt", "w")
            #fileout4 = open("Srbd_v2_aa_list.txt", "w")
            #fileout5 = open("Spbs_nt_list.txt", "w")
            #fileout6 = open("Spbs_aa_list.txt", "w")

            #generating workbook in parallel
            all_wbname = args.run_folder+"/results/"+runID+"_AllVariantDetails.xlsx"
            all_wb = Workbook()
            ws_Srbd_v2 = all_wb.active
            ws_Srbd_v2.title = "Srbd_v2"
            ws_spbs = all_wb.create_sheet(title="Spbs")
            ws_rdrp = all_wb.create_sheet(title="RdRP")
            ws_list = [ws_Srbd_v2, ws_spbs, ws_rdrp]

            amps = ["Srbd_v2", "Spbs", "RdRP"]
            header = []

            for amp in amps:
                header = ["sample","well",amp+"_total_read#",amp+"_top1_read#","nt_var","aa_var","present_in_sec_hit","present_in_third_hit","filename"]
                i = amps.index(amp)
                ws = ws_list[i]
                ws.append(header)

        for filename in input_file_set:
            #Part-1 binning sequences and creating "sequence - readcount" file

            #Using regex to parse info
            matches = re.match(r'(.*?)_(.*?)_S.{1,4}_.*?_(.{2,3})_.*', filename)
            date = matches.group(1)
            sampleID = matches.group(2)
            well = matches.group(3)

            #include well if duplicate sampleID found
            if sampleID in sampleID_dups:
                sampleID = sampleID+"_"+well

            long_read_dict = {}
            it=2
            filenamepath = args.run_folder+"/R1_files/"+filename

            with open(filenamepath) as fastq:
                for line in fastq:
                    it+=1

                    if it%4==0:
                        rseq = line.strip()
                        adap_pos = rseq.find("AGATCGGAAG")
                        seq = rseq[:adap_pos]

                        if len(seq)>90:
                            if seq not in long_read_dict:
                                long_read_dict[seq]=1
                            else:
                                long_read_dict[seq]+=1

            fastq.close()

            #Part-2 binning sequences with read counts into PCR pools and sorting writing top

            rdrp_ = []
            Srbd_v2_ = []
            Spbs_ = []

            rdrp_c_ = 0.001
            Srbd_v2_c_ = 0.001
            Spbs_c_ = 0.001

            N501Yc = 0
            E484Kc = 0
            S494Pc = 0
            P681Hc = 0
            P681Rc = 0
            T478Kc = 0

            Srbd_v2_WTc = 0
            Spbs_WTc = 0

            #fileout1 = open((filename[:-4] + '_RdRP_count.txt'), 'w')
            #fileout4 = open((filename[:-4] + '_S_RBD_count.txt'), 'w')
            #fileout8 = open((filename[:-4] + '_S_PBS_count.txt'), 'w')

            long_read_t = long_read_dict.items()

            for key,val in long_read_t:
                if "ACTGCTTATG" in key and "CGATAAGT" in key :
                    rdrp_.append((val,key))
                    rdrp_c_ = rdrp_c_ + val

                #elif "AGAGATATTT" in key: #select_var target
                elif "AGAGAGATATTT" in key: #all_top target
                    Srbd_v2_.append((val, key))
                    Srbd_v2_c_ = Srbd_v2_c_+ val

                    if run_select_vars:
                        if "CCCACTTATG" in key:
                            N501Yc = N501Yc + val

                        if "TGTTAAAGGT" in key:
                            E484Kc= E484Kc + val

                        if "ACAACCATAT" in key:
                            S494Pc= S494Pc + val

                        if "AGCAAACCTT" in key:
                            T478Kc = T478Kc + val

                        if key in Srbd_v2_amp:
                            Srbd_v2_WTc = Srbd_v2_WTc + val

                elif "TCAGACTCAGAC" in key and "GTGCAGAA" in key:
                    Spbs_.append((val, key))
                    Spbs_c_ = Spbs_c_+ val

                    if run_select_vars:
                        if "TCTCATCGG" in key:
                            P681Hc=P681Hc+val
                        if "TCTCGTCGG" in key:
                            P681Rc=P681Rc+val
                        if key in Spbs_amp:
                            Spbs_WTc=Spbs_WTc+val

            #build comb_mutdic
            if run_select_vars:

                comb_mutdic = {}
                comb_mutdic["sample"]=sampleID
                #if [amp]_c_ values not set to 0.001 will throw divide by zero error
                comb_mutdic["Srbd_v2_WT_percentage"]=("{:.2f}".format(Srbd_v2_WTc/Srbd_v2_c_*100))
                comb_mutdic["Spbs_WT_percentage"]=("{:.2f}".format(Spbs_WTc/Spbs_c_*100))
                comb_mutdic["total_N501Y_percentage"]=("{:.2f}".format(N501Yc/Srbd_v2_c_*100))
                comb_mutdic["total_E484K_percentage"]=("{:.2f}".format(E484Kc/Srbd_v2_c_*100))
                comb_mutdic["total_S494P_percentage"]=("{:.2f}".format(S494Pc/Srbd_v2_c_*100))
                comb_mutdic["total_P681H_percentage"]=("{:.2f}".format(P681Hc/Spbs_c_*100))
                comb_mutdic["total_P681R_percentage"]=("{:.2f}".format(P681Rc/Spbs_c_*100))
                comb_mutdic["total_T478K_percentage"]=("{:.2f}".format(T478Kc/Srbd_v2_c_*100))
                comb_mutdic["Spbs_total_read_count"]=Spbs_c_
                comb_mutdic["Srbd_v2_total_read_count"]=Srbd_v2_c_

            #build aa_dict and amp nt/aa_dicts
            if run_all_top:

                aa_dic = {}
                codonaatable = args.resource_folder+'/codon_aa_table.csv'
                with open(codonaatable) as condonf:
                    for line in condonf:
                        info = line.strip()
                        infol = info.split(",")
                        aa_dic[(infol[0])] = infol[1]
                condonf.close()

                ##internal fxns: build_ntdict(amp, nt_dict, nt_num, it) and build_aadict(aa_dict, amp_dict, amp, pos, aa_num)
                RdRP_nt_dic = {}
                RdRP_aa_dic = {}
                RdRP_nt_dic = build_ntdict(RdRP_amp, RdRP_nt_dic, 15490, 0)
                RdRP_aa_dic = build_aadict(aa_dic, RdRP_aa_dic, RdRP_amp, 0, 684) #adjusted from 575 to 684 to match ref

                Srbd_v2_nt_dic = {}
                Srbd_v2_aa_dic = {}
                Srbd_v2_nt_dic = build_ntdict(Srbd_v2_amp, Srbd_v2_nt_dic, 22948, 0)
                Srbd_v2_aa_dic = build_aadict(aa_dic, Srbd_v2_aa_dic, Srbd_v2_amp, 1, 463)

                Spbs_nt_dic = {}
                Spbs_aa_dic = {}
                Spbs_nt_dic = build_ntdict(Spbs_amp, Spbs_nt_dic, 23571, 0)
                Spbs_aa_dic = build_aadict(aa_dic, Spbs_aa_dic, Spbs_amp, 2, 671)

            #Rdrp
            rdrp_.sort(reverse=True)
            rdrp_ltop = rdrp_[:1]
            var_cond=0

            for val,item in rdrp_ltop:

                if run_select_vars:
                    if item[:96] not in RdRP_amp:
                        if val>2:
                            RdRP_PCR_file.write(">" + sampleID + "_RdRP_" + str(val) + "of" + str(rdrp_c_) + "\n" + item + "\n")

                    #if item in "GATGCCACAACTGCTTATGCTAATAGTGTTTTTAACATTTGTCAAGCTGTCACGGCCAATGTTAATGCACTTTTATCTACTGATGGTAACAAAATTGCCGATAAGTATGTCCGCAA":
                        #fileoutMutSum.write(filename + ",WT," + str(val) + ",")
                    #else : fileoutMutSum.write(filename + ",new variant," + str(val) + ",")

                #for val,item in rdrp_:
                    #fileout1.write(">" + filename[:18] + "_RdRP_" + str(val) + "of" + str(rdrp_c_) + " " + item + "\n")
                    #fileout1.write(">" + filename[:18] + "_RdRP_" + str("{:.2f}".format(val/rdrp_c_*100)) + "%" + " " + item + "\n")

                if run_all_top:
                    ws_data = []

                    if val>3 :
                        seq=item[2:96]
                        nt_num = 15492

                        for nt in seq :
                            if nt_num<15587 and RdRP_nt_dic[nt_num]!= nt :
                                RdRP_Var_file.write(sampleID + "," + well +","+ str(rdrp_c_) + "," + str(val) + "," + RdRP_nt_dic[nt_num]+str(nt_num)+nt+",")
                                ws_data = [sampleID, well, str(rdrp_c_),str(val) , RdRP_nt_dic[nt_num]+str(nt_num)+nt]

                                var_aa_num = 684+int((nt_num-15490)/3) #adjusted from 575 to 684 to match ref
                                var_aa_start = ((var_aa_num-684)*3)
                                var_aa_end = var_aa_start+3
                                ncodon = seq[var_aa_start:var_aa_end]

                                if len(ncodon)==3 and "N" not in ncodon:
                                    RdRP_Var_file.write(RdRP_aa_dic[var_aa_num]+str(var_aa_num)+aa_dic[ncodon]+",")
                                    ws_data.append(RdRP_aa_dic[var_aa_num]+str(var_aa_num)+aa_dic[ncodon])

                                else:
                                    RdRP_Var_file.write(RdRP_aa_dic[var_aa_num]+str(var_aa_num)+"NA"+",")
                                    ws_data.append(RdRP_aa_dic[var_aa_num]+str(var_aa_num)+"NA")

                                for val2,item2 in rdrp_[1:2] :
                                    seq2=item2[2:96]

                                    if ncodon == seq2[var_aa_start:var_aa_end]:
                                        RdRP_Var_file.write("yes#" + str(val2) + ",")
                                        ws_data.append("yes#" + str(val2))
                                    else:
                                        RdRP_Var_file.write("no#" + str(val2) + ",")
                                        ws_data.append("no#" + str(val2))

                                for val3,item3 in rdrp_[2:3] :
                                    seq3=item3[2:96]

                                    if ncodon == seq3[var_aa_start:var_aa_end]:
                                        RdRP_Var_file.write("yes#" + str(val3) + "," + filename + "\n")
                                        ws_data.extend(["yes#" + str(val3),filename])
                                        ws_rdrp.append(ws_data)

                                    else:
                                        RdRP_Var_file.write("no#" + str(val3) + "," + filename + "\n")
                                        ws_data.extend(["no#" + str(val3), filename])
                                        ws_rdrp.append(ws_data)

                                if len(ws_data) < len(header):
                                    numdatamissing = len(header)-len(ws_data)
                                    print("\t\t\t!!! Warning:", sampleID, "is missing ", str(numdatamissing), "data value(s) in file: SparSeq_RdRP_top1_Var_summary_table.txt !!!")
                                    missingdata = "MISSING," * numdatamissing
                                    RdRP_Var_file.write(missingdata+filename+"\n")
                                    ws_data.extend([("MISSING" * numdatamissing), filename])
                                    ws_rdrp.append(ws_data)

                                var_cond =var_cond+1
                            nt_num=nt_num+1

                        if var_cond==0 :
                            RdRP_Var_file.write(sampleID + "," + well +","+ str(rdrp_c_) + "," + str(val) + ",WT,WT,NA,NA," + filename + "\n")
                            ws_data = [sampleID, well, str(rdrp_c_),str(val),"WT","WT","NA","NA",filename]
                            ws_rdrp.append(ws_data)

                    else:
                        RdRP_Var_file.write(sampleID + "," + well +","+ str(rdrp_c_) + "," + str(val) + ",low,low,low,low," + filename + "\n")
                        ws_data = [sampleID, well, str(rdrp_c_),str(val),"low","low","low","low",filename]
                        ws_rdrp.append(ws_data)


            #Srbd_v2
            Srbd_v2_.sort(reverse=True)
            Srbd_v2_ltop = Srbd_v2_[:1]

            for val,item in Srbd_v2_ltop:

                if run_select_vars:
                    if item[:129] in Srbd_v2_amp:
                        comb_mutdic["Srbd_v2_tophit"]="WT"
                        comb_mutdic["Srbd_v2_tophit_percentage"] = ("{:.2f}".format(val/Srbd_v2_c_*100))

                    elif item[:129] not in Srbd_v2_amp:

                        if val>2:
                            Srbd_v2_PCR_file.write(">"+sampleID + "_Srbd_v2_" + str(val) + "of" + str(rdrp_c_) + "\n" + item + "\n")
                        comb_mutdic["Srbd_v2_tophit"]="-"

                        if "CCCACTTATG" in item:
                            comb_mutdic["Srbd_v2_tophit"]=comb_mutdic["Srbd_v2_tophit"] + "N501Y-"
                            comb_mutdic["Srbd_v2_tophit_percentage"] = ("{:.2f}".format(val/Srbd_v2_c_*100))

                        if "TGTTAAAGGT" in item:
                            comb_mutdic["Srbd_v2_tophit"]=comb_mutdic["Srbd_v2_tophit"] + "E484K-"
                            comb_mutdic["Srbd_v2_tophit_percentage"] = ("{:.2f}".format(val/Srbd_v2_c_*100))

                        if "ACAACCATAT" in item:
                            comb_mutdic["Srbd_v2_tophit"]=comb_mutdic["Srbd_v2_tophit"] + "S494P-"
                            comb_mutdic["Srbd_v2_tophit_percentage"] = ("{:.2f}".format(val/Srbd_v2_c_*100))

                        if "AGCAAACCTT" in item:
                                comb_mutdic["Srbd_v2_tophit"]=comb_mutdic["Srbd_v2_tophit"] + "T478K-"
                                comb_mutdic["Srbd_v2_tophit_percentage"] = ("{:.2f}".format(val/Srbd_v2_c_*100))

                        if  "ACAACCATAT" not in item and "TGTTAAAGGT" not in item and "CCCACTTATG" not in item and "AGCAAACCTT" not in item:
                            comb_mutdic["Srbd_v2_tophit"]=comb_mutdic["Srbd_v2_tophit"] + "other variants-"
                            comb_mutdic["Srbd_v2_tophit_percentage"] = ("{:.2f}".format(val/Srbd_v2_c_*100))

                    #for val,item in Srbd_v2_ :
                        #fileout4.write(">" + filename[:18] + "_Srbd_v2_" + str(val) + "of" + str(rdrp_c_) + " " + item + "\n")

                if run_all_top:
                    ws_data = []

                    if val>3 :
                        seq=item[1:127]
                        nt_num = 22949
                        var_cond=0

                        for nt in seq :

                            if nt_num<23075 and Srbd_v2_nt_dic[nt_num]!= nt :
                                Srbd_v2_Var_file.write(sampleID + "," + well +","+ str(Srbd_v2_c_) + "," + str(val) + "," + Srbd_v2_nt_dic[nt_num]+str(nt_num)+nt+",")
                                ws_data = [sampleID, well, str(Srbd_v2_c_),str(val),Srbd_v2_nt_dic[nt_num]+str(nt_num)+nt]


                                var_aa_num = 463+int((nt_num-22949)/3)
                                var_aa_start = ((var_aa_num-463)*3)
                                var_aa_end = var_aa_start+3
                                ncodon = seq[var_aa_start:var_aa_end]

                                if len(ncodon)==3 and "N" not in ncodon:
                                    Srbd_v2_Var_file.write(Srbd_v2_aa_dic[var_aa_num]+str(var_aa_num)+aa_dic[ncodon]+",")
                                    ws_data.append(Srbd_v2_aa_dic[var_aa_num]+str(var_aa_num)+aa_dic[ncodon])

                                else:
                                    Srbd_v2_Var_file.write(Srbd_v2_aa_dic[var_aa_num]+str(var_aa_num)+"NA"+",")
                                    ws_data.append(Srbd_v2_aa_dic[var_aa_num]+str(var_aa_num)+"NA")

                                for val2,item2 in Srbd_v2_[1:2] :
                                    seq2=item2[1:127]
                                    if ncodon == seq2[var_aa_start:var_aa_end]:
                                        Srbd_v2_Var_file.write("yes#" + str(val2) + ",")
                                        ws_data.append("yes#" + str(val2))

                                    else:
                                        Srbd_v2_Var_file.write("no#" + str(val2) + ",")
                                        ws_data.append("no#" + str(val2))

                                for val3,item3 in Srbd_v2_[2:3] :
                                    seq3=item3[1:127]
                                    if ncodon == seq3[var_aa_start:var_aa_end]:
                                        Srbd_v2_Var_file.write("yes#" + str(val3) + "," + filename + "\n")
                                        ws_data.extend(["yes#" + str(val3), filename])
                                        ws_Srbd_v2.append(ws_data)

                                    else:
                                        Srbd_v2_Var_file.write("no#" + str(val3) + "," + filename + "\n")
                                        ws_data.extend(["no#" + str(val3), filename])
                                        ws_Srbd_v2.append(ws_data)

                                if len(ws_data) < len(header):
                                    numdatamissing = (len(header)-len(ws_data))-1 #-1 for filename
                                    print("\t\t\t!!! Warning:", sampleID, "is missing", str(numdatamissing), "data value(s) in file: SparSeq_Srbd_v2_top1_Var_summary_table.txt !!!")
                                    missingdata = "MISSING," * numdatamissing
                                    Srbd_v2_Var_file.write(missingdata+filename+"\n")
                                    ws_data.extend([("MISSING" * numdatamissing), filename])
                                    ws_Srbd_v2.append(ws_data)

                                var_cond =var_cond+1
                            nt_num=nt_num+1

                        if var_cond==0:
                            Srbd_v2_Var_file.write(sampleID + "," + well +","+ str(Srbd_v2_c_) + "," + str(val) + ",WT,WT,NA,NA," + filename + "\n")
                            ws_data = [sampleID, well, str(Srbd_v2_c_), str(val),"WT","WT","NA","NA",filename]
                            ws_Srbd_v2.append(ws_data)

                    else:
                        Srbd_v2_Var_file.write(sampleID + "," + well +","+ str(Srbd_v2_c_) + "," + str(val) + ",low,low,low,low," + filename + "\n")
                        ws_data = [sampleID, well, str(Srbd_v2_c_),str(val),"low","low","low","low",filename]
                        ws_Srbd_v2.append(ws_data)

            #Spbs
            Spbs_.sort(reverse=True)
            Spbs_ltop = Spbs_[:1]

            for val,item in Spbs_ltop:

                if run_select_vars:
                    if item not in Spbs_amp:
                        if val>2 : Spbs_PCR_file.write(">"+ sampleID + "_Spbs_" + str(val) + "of" + str(rdrp_c_) + "\n" + item + "\n")

                    if item[:87] in Spbs_amp:
                        comb_mutdic["Spbs_tophit"]="WT"
                        comb_mutdic["Spbs_tophit_percentage"] = ("{:.2f}".format(val/Spbs_c_*100))

                    elif "TCTCATCGG" in item:
                        comb_mutdic["Spbs_tophit"]="-P681H-"
                        comb_mutdic["Spbs_tophit_percentage"] = ("{:.2f}".format(val/Spbs_c_*100))

                    elif "TCTCGTCGG" in item:
                        comb_mutdic["Spbs_tophit"]="-P681R-"
                        comb_mutdic["Spbs_tophit_percentage"] = ("{:.2f}".format(val/Spbs_c_*100))

                    else:
                        comb_mutdic["Spbs_tophit"]="-other variants-"
                        comb_mutdic["Spbs_tophit_percentage"] = ("{:.2f}".format(val/Spbs_c_*100))

                if run_all_top:
                    ws_data = []

                    if val>3 :
                        seq=item[2:87]
                        nt_num=23573
                        var_cond=0

                        for nt in seq :
                            if nt_num<23662 and Spbs_nt_dic[nt_num] != nt :
                                Spbs_Var_file.write(sampleID + "," + well + ","+ str(Spbs_c_) + "," + str(val) + ","+Spbs_nt_dic[nt_num]+str(nt_num)+nt+",")
                                ws_data = [sampleID, well, str(Spbs_c_), str(val), Spbs_nt_dic[nt_num]+str(nt_num)+nt]

                                var_aa_num = 671+int((nt_num-23573)/3)
                                var_aa_start = ((var_aa_num-671)*3)
                                var_aa_end = var_aa_start+3
                                ncodon = seq[var_aa_start:var_aa_end]

                                if len(ncodon)==3 and "N" not in ncodon:
                                    Spbs_Var_file.write(Spbs_aa_dic[var_aa_num]+str(var_aa_num)+aa_dic[ncodon]+",")
                                    ws_data.append(Spbs_aa_dic[var_aa_num]+str(var_aa_num)+aa_dic[ncodon])

                                else:
                                    Spbs_Var_file.write(Spbs_aa_dic[var_aa_num]+str(var_aa_num)+"NA"+",")
                                    ws_data.append(Spbs_aa_dic[var_aa_num]+str(var_aa_num)+"NA")

                                for val2,item2 in Spbs_[1:2] :
                                    seq2=item2[2:87]
                                    if ncodon == seq2[var_aa_start:var_aa_end]:
                                        Spbs_Var_file.write("yes#" + str(val2) + ",")
                                        ws_data.append("yes#" + str(val2))

                                    else:
                                        Spbs_Var_file.write("no#" + str(val2) + ",")
                                        ws_data.append("no#" + str(val2))

                                for val3,item3 in Spbs_[2:3] :
                                    seq3=item3[2:87]
                                    if ncodon == seq3[var_aa_start:var_aa_end]:
                                        Spbs_Var_file.write("yes#" + str(val3) + "," + filename + "\n")
                                        ws_data.extend(["yes#" + str(val3),filename])
                                        ws_spbs.append(ws_data)

                                    else:
                                        Spbs_Var_file.write("no#" + str(val3) + "," + filename + "\n")
                                        ws_data.extend(["no#" + str(val3),filename])
                                        ws_spbs.append(ws_data)


                                if len(ws_data) < len(header):
                                    numdatamissing = len(header)-len(ws_data)
                                    print("\t\t\t!!! Warning:", sampleID, "is missing", str(numdatamissing), "data value(s) in file: SparSeq_Spbs_top1_Var_summary_table.txt !!!")
                                    missingdata = "MISSING," * numdatamissing
                                    Spbs_Var_file.write(missingdata+filename+"\n")
                                    ws_data.extend([("MISSING" * numdatamissing), filename])
                                    ws_spbs.append(ws_data)

                                var_cond+=1
                            nt_num+=1

                        if var_cond==0 :
                            Spbs_Var_file.write(sampleID + "," + well +","+ str(Spbs_c_) + "," + str(val) + ",WT,WT,NA,NA," + filename + "\n")
                            ws_data = [sampleID, well, str(Spbs_c_),str(val),"WT","WT","NA","NA",filename]
                            ws_spbs.append(ws_data)

                    else:
                        Spbs_Var_file.write(sampleID + "," + well +","+ str(Spbs_c_) + "," + str(val) + ",low,low,low,low," + filename + "\n")
                        ws_data = [sampleID, well, str(Spbs_c_),str(val),"low","low","low","low",filename]
                        ws_spbs.append(ws_data)

            if run_select_vars:
                ws_seldata = []

                for items in second_info_l:

                    if items in comb_mutdic:
                        fileout_combMut.write(str(comb_mutdic[items]) + ",")
                        selvarstable.write(str(comb_mutdic[items]) + ",")
                        ws_seldata.append(str(comb_mutdic[items]))

                    else:
                        fileout_combMut.write( "NA,")
                        selvarstable.write( "NA,")
                        ws_seldata.append("NA")

                fileout_combMut.write(filename+"\n")
                selvarstable.write(filename+"\n")
                ws_seldata.append(filename)
                ws_selvars.append(ws_seldata)

                #for val,item in Spbs_ :
                    #fileout8.write(">" + filename[:18] + "_Spbs_" + str(val) + "of" + str(rdrp_c_) + " " + item + "\n")
                #fileoutMutSum.write(filename + ",N501Y," + str(N501Yc) +","+ str(Srbd_v2_c_)+ ",Srbd_v2" +"\n")
                #fileoutMutSum.write(filename + ",E484K," + str(E484Kc) +","+ str(Srbd_v2_c_)+ ",Srbd_v2" +"\n")
                #fileoutMutSum.write(filename + ",S494P," + str(S494Pc) +","+ str(Srbd_v2_c_)+ ",Srbd_v2" +"\n")
                #fileoutMutSum.write(filename + ",P681H," + str(P681Hc) +","+ str(Spbs_c_)+ ",Spbs" +"\n")

        if run_select_vars:
            Srbd_v2_PCR_file.close()
            Spbs_PCR_file.close()
            RdRP_PCR_file.close()
            fileout_combMut.close()
            selvarstable.close()
            selvars_wb.save(filename = selvars_wbname) #save wb
            print("\t\t\t> select_vars now complete...")

        if run_all_top:
            Spbs_Var_file.close()
            RdRP_Var_file.close()
            Srbd_v2_Var_file.close()
            all_wb.save(filename = all_wbname) #save wb
            print("\t\t\t> all_top now complete...")

    if run_bowtie_count:

        #Not assuming count files will be the same as fastqs so resetting
        sampleID_list = [] #used to check for duplicate sampleIDs
        input_file_set = []

        for samcountout in os.listdir(args.run_folder+"/R1_files/samcounts"):
            if ".count.txt" in samcountout:
                matches = re.match(r'.*?_(.*?)_.*?', samcountout)
                sampleID_list.append(matches.group(1))
                input_file_set.append(samcountout)

        input_file_set.sort()

        sID = set()
        sampleID_dups = {x for x in sampleID_list if x in sID or (sID.add(x) or False)} #lists duplicate sampleID

        #defining outputs
        bctable_file = open(args.run_folder+"/results/bowtie_count_table_"+version+".txt", "w") #fileout
        bctable_file.write("samples")
        ctable_file = open(args.run_folder+"/"+runID+"_CountTable.txt", "w") #ctableout
        ctable_file.write("sample")

        #generating workbook in parallel
        counts_wbname = args.run_folder+"/results/"+runID+"_CountTable.xlsx"
        counts_wb = Workbook()
        ws_counts = counts_wb.active
        ws_counts.title = "bowtie_count_table_"+version

        it=0
        remove_data = ["Srbd", "__no_feature", "__ambiguous", "__too_low_aQual","__alignment_not_unique"]
        count_header = ["sample"]

        for countfile in input_file_set:
            it+=1
            countfilepath = args.run_folder+"/R1_files/samcounts/"+countfile

            #Using regex to parse sampleID; include well to account for duplicates e.g. 6062021_H2O-multi_S41_R2C3_23M_S41_R1_001.fastq
            matches = re.match(r'(.*?)_(.*?)_S.{1,4}_.*?_(.{2,3})_.*', countfile)
            date = matches.group(1)
            sampleID = matches.group(2)
            well = matches.group(3)

            if sampleID in sampleID_dups:
                sampleID = sampleID+"_"+well

            count_data = [sampleID]
            with open(countfilepath) as f1:

                header= ""
                data = sampleID
                totalrawreads = 0
                totalviral = 0

                for line in f1:
                    info = line.strip()
                    infol = info.split()

                    if infol[0] not in remove_data:

                        if it==1: #if first iteration add header line
                            header+=","+infol[0]
                            count_header.append(infol[0])

                        data+=","+infol[1]
                        count_data.append(int(infol[1]))

                        if infol[0]=='Rdrp' or infol[0]=='Spoly' or infol[0]=='Srbd_v2':
                            totalviral+=int(infol[1])

                    totalrawreads+=int(infol[1]) #include unwanted fields in count

                if it == 1:
                    bctable_file.write(header+"\n")
                    header+=",total.viral,total.raw.reads,filename\n"
                    ctable_file.write(header)
                    count_header.extend(["total.viral", "total.raw.reads", "filename"])
                    ws_counts.append(count_header)

            bctable_file.write(data+","+countfile+"\n")
            data+=","+str(totalviral)+","+str(totalrawreads)+","+countfile+"\n"
            ctable_file.write(data)
            count_data.extend([totalviral, totalrawreads, countfile])
            ws_counts.append(count_data)

        f1.close()
        bctable_file.close()
        ctable_file.close()
        counts_wb.save(filename = counts_wbname) #save wb

        #clean up
        if not args.intermediates:
            try:
                for f in os.listdir(run_folder+"/R1_files/samcounts"):
                    if  re.search("fastq.sam.count.txt", f):
                        os.remove(os.path.join(run_folder+"/R1_files/samcounts/", f))
            except:
                pass

        print("\t\t\t> bowtie_count now complete...")

    if run_variant_agg:
        #variant agg

        import pandas
        import datetime
        pandas.options.mode.chained_assignment = None #ignore copy warnings

        #creating dfs
        countstable = args.run_folder+"/"+runID+"_CountTable.txt"
        countdf = pandas.read_csv(countstable)  #still reading in file to account for start argument
        
        Srbd_v2file = args.run_folder+"/results/SparSeq_Srbd_v2_top1_Var_summary_table.txt"
        Srbd_v2df = pandas.read_csv(Srbd_v2file, names=("sample", "Srbd_v2_total_read.", "Srbd_v2_top1_read.",   "aa_var"), usecols=[0,2,3,5], header=0) #only loading relative information
        Srbd_v2df = Srbd_v2df.groupby('sample', as_index=False).agg({'Srbd_v2_total_read.':'first', 'Srbd_v2_top1_read.':'first', 'aa_var':"; ".join}) #agg data
        Srbd_v2df.drop_duplicates()

        combined = countdf.merge(Srbd_v2df, how="left", on='sample')
        combined.loc[:,'type'] = combined['sample'].apply(lambda x: 'Sample' if (x.startswith('X') | x.startswith('W') | x.startswith('Y') | x.startswith('Z') | x.startswith('Neg')) else 'Control')


        controlsOnly = combined[combined['type'] == 'Control']
        samplesOnly = combined[combined['type'] =='Sample']
        controlsOnly = controlsOnly[["sample", "ACTB", "ACTG", "Rdrp",  "Spoly", "Srbd_v2"]]

        #generating median of Srbd_v2 control values
        Srbd_v2_median = controlsOnly["Srbd_v2"].median()

        #generating absolute difference between each Srbd_v2 value and the overall median value
        controlsOnly.loc[:,"medianDiff"] = controlsOnly["Srbd_v2"].apply(lambda x: abs(x-Srbd_v2_median))

        #generating MAD by then take the median of medianDiff
        Srbd_v2_diff_median = controlsOnly["medianDiff"].median()

        #-use original median of the controls + (2xMAD)
        #-that value is now the threshold for pass or fail for the QC check
        combinedmedians = Srbd_v2_median + (2*Srbd_v2_diff_median)
        if combinedmedians < 10 : combinedmedians = 10 #min threshold

        vardetfile = args.run_folder+"/"+runID+"_SelectedVariants.txt"
        vardet = pandas.read_csv(vardetfile, usecols=[0,1,2,3,7,4,5,6], header=0) #only loading relative information
        vardet.loc[:,"Srbd_v2_diff"] = abs(vardet["Srbd_v2_WT_percentage"]-vardet["Srbd_v2_tophit_percentage"])

        # marking rows with a difference > 15 between tophit% and wt%
        Srbd_v2TopHitWarnings = vardet[(vardet["Srbd_v2_diff"] < 15) & (vardet["Srbd_v2_tophit"] != "WT")]

        today = datetime.datetime.strftime(datetime.datetime.now(), "%B %d, %Y")
        samplesOnly.loc[:, "Date"] = "Processed "+str(today)

        samplesOnly.loc[:,"E484K"] = samplesOnly["aa_var"].apply(lambda x: "+" if ("Glu484Lys" in str(x)) else "-")
        samplesOnly.loc[:,"N501Y"] = samplesOnly["aa_var"].apply(lambda x: "+" if ("Asn501Tyr" in str(x)) else "-")
        samplesOnly.loc[:,"QC_Check_MinReads"] = (samplesOnly["Srbd_v2"].apply(lambda x: "Pass" if (x >= combinedmedians) else "Fail"))

        #check list of fails
        samplesOnly.loc[:,"QC_Check_N501Y_E484K_Coverage"] = samplesOnly["sample"].isin(Srbd_v2TopHitWarnings["sample"])
        samplesOnly.loc[:,"QC_Check_N501Y_E484K_Coverage"] = samplesOnly["QC_Check_N501Y_E484K_Coverage"].apply(lambda x: "Fail" if x =="True" else "Pass") #reassign values
        samplesOnly.loc[:,"QC_Check_raw"] = samplesOnly["QC_Check_MinReads"]+"_"+samplesOnly["QC_Check_N501Y_E484K_Coverage"]
        samplesOnly.loc[:,"QC_Check"] = samplesOnly["QC_Check_raw"].apply(lambda x: "Pass" if x == "Pass_Pass"  else "Indeterminate")

        finalReport = samplesOnly[["sample", "E484K", "N501Y", "QC_Check_MinReads", "QC_Check_N501Y_E484K_Coverage", "QC_Check", "Date"]]

        finalreportfile = args.run_folder+"/results/sparseq_report_"+runID+".xlsx"
        finalReport.to_excel(finalreportfile, sheet_name=runID,index=False)
        print("\t\t\t> variant_agg now complete...")

#Defining internal functions:
def build_ntdict(amp, nt_dict, nt_num, it):
    amp = amp
    amp_nt_dict = nt_dict
    amp_nt_num = nt_num
    it = it

    for nt in amp:
        amp_nt_dict[amp_nt_num]=amp[it]
        amp_nt_num+=1
        it+=1

    return amp_nt_dict

def build_aadict(aa_dict, amp_dict, amp, pos, aa_num):
    codon_aa_dict = aa_dict
    amp_aa_dict = aa_dict
    amp = amp
    amp_pos = pos
    amp_aa_num = aa_num

    for aa in amp:
        if amp_pos+3 < len(amp):
            codon = amp[amp_pos:(amp_pos+3)]
            amp_aa_dict[amp_aa_num] = codon_aa_dict[codon]
            amp_pos+=3
            amp_aa_num+=1
        else : break

    return amp_aa_dict

main()
